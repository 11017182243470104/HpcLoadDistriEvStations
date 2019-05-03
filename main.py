# main file with procedure
from random_generator import *
from opt_model_pwl import *
import openpyxl

# input of the samplesize by the user
print("How big should the sample be?")
samplesize = input()

#  Creating the Excel file
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Multi-CPO-Setting"
ws1["A1"] = "Run"
ws1["B1"] = "Opt objective value "
ws1["C1"] = "Opt average"
ws1["D1"] = "Opt max"
ws1["E1"] = "Opt min"
ws1["F1"] = "Smpl objective value"
ws1["G1"] = "Smpl average"
ws1["H1"] = "Smpl max"
ws1["I1"] = "Smpl min"
ws1["J1"] = "Opt objective val w real exp"
ws1["K1"] = "Opt average w real exp"
ws1["L1"] = "Opt max w real exp"
ws1["M1"] = "Opt min w real exp"
ws1["N1"] = "Smpl objective val w real exp"
ws1["O1"] = "Smpl average w real exp"
ws1["P1"] = "Smpl max w real exp"
ws1["Q1"] = "Smpl min w real exp"

# Filling the Excel file with data
count = 0
for run in range(int(samplesize)):
    solution = solve_instance(create_rand_instance())
    print("optimized: ", solution[0])
    print("simple: ", solution[1])

    ws1.cell(column=1, row=run + 2, value=run + 1)
    for k in range(4):
        for i in range(4):
            ws1.cell(column=2 + i + (k * 4), row=run + 2, value=solution[k][i])

    if solution[0][0] < solution[1][0]:
        count += 1

wb.save("Output_80.xlsx")

# Output of the samplesize and the optimization rate
print("\n%d random instances calculated! \r" % int(samplesize))
print("Percentage model is better than simple algorithm: %.2f%%" % (count / int(samplesize) * 100))
print("Variance, average, maximum & minimum for simple and optimized model have been save to Output.xlsx")
